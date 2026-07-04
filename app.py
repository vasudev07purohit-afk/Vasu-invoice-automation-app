import os
import tempfile
import streamlit as st
from openpyxl import load_workbook

import invoice_automation as engine

st.set_page_config(page_title="Invoice → VAT Ledger Automation", page_icon="📄", layout="centered")

st.title("📄 Invoice → VAT Ledger Automation")
st.write(
    "Upload your invoices ZIP and your Excel VAT-summary workbook. "
    "The tool reads each invoice PDF, matches it to the right ledger row, "
    "and gives you back a new Excel file with the details filled in — "
    "your original file is never changed."
)

st.divider()

col1, col2 = st.columns(2)
with col1:
    zip_file = st.file_uploader("1. Invoices ZIP", type=["zip"])
with col2:
    excel_file = st.file_uploader("2. Excel workbook", type=["xlsx"])

run_clicked = st.button("Run matching", type="primary", disabled=not (zip_file and excel_file))

if run_clicked:
    with st.spinner("Reading invoices and matching to your ledger — this can take a minute..."):
        with tempfile.TemporaryDirectory() as tmp:
            zip_path = os.path.join(tmp, "invoices.zip")
            excel_path = os.path.join(tmp, "input.xlsx")
            out_path = os.path.join(tmp, "output.xlsx")

            with open(zip_path, "wb") as f:
                f.write(zip_file.getbuffer())
            with open(excel_path, "wb") as f:
                f.write(excel_file.getbuffer())

            try:
                engine.run(zip_path, excel_path, out_path)
            except Exception as e:
                st.error(f"Something went wrong while processing: {e}")
                st.stop()

            with open(out_path, "rb") as f:
                result_bytes = f.read()

            # Pull a quick summary from the Reconciliation sheet for on-screen display
            wb = load_workbook(out_path)
            rc = wb["Reconciliation"]
            n_ok = n_mismatch = n_missing = n_unused = 0
            for row in rc.iter_rows(min_row=2, values_only=True):
                status = row[3] if len(row) > 3 else None
                if status == "OK":
                    n_ok += 1
                elif status == "AMOUNT MISMATCH":
                    n_mismatch += 1
                elif status == "NO PDF FOUND":
                    n_missing += 1
                elif status == "UNUSED PDF":
                    n_unused += 1

    st.success("Done! Here's a quick summary:")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("🟩 Matched OK", n_ok)
    m2.metric("🟨 Amount mismatch", n_mismatch)
    m3.metric("🟥 No PDF found", n_missing)
    m4.metric("⬜ Unused PDFs", n_unused)

    st.download_button(
        "⬇️ Download completed workbook",
        data=result_bytes,
        file_name="Upload_WITH_INVOICE_DETAILS.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    st.caption(
        "Open the 'Reconciliation' sheet in the downloaded file for the full "
        "row-by-row breakdown, colour-coded the same way as above."
    )

st.divider()
with st.expander("How matching works"):
    st.markdown(
        """
- Reads the invoice reference out of each ledger row's **Particulars** text.
- Finds the matching invoice PDF in the ZIP, cross-checked against the supplier name
  so numbers can't accidentally match the wrong supplier.
- Fills **Basic Amt / VAT / Invoice Value** only where those cells are blank —
  it never overwrites figures you've already entered.
- Sets **Invoice (Y/N) = Y** on matched rows.
- Adds detail columns: PDF File, Inv No (PDF), Inv Date (PDF), PDF Net/VAT/Gross,
  Item Details (PDF).
- Builds a colour-coded **Reconciliation** sheet: green = matched & amounts agree,
  yellow = matched but amount differs, red = ledger row with no invoice found,
  grey = invoice in the ZIP never used.
        """
    )
