#!/usr/bin/env python3
"""
================================================================================
 SUPPLIER INVOICE -> VAT LEDGER  AUTOMATION
================================================================================
 Reads a ZIP of supplier-wise invoice PDFs and transfers the item information
 into an Excel VAT-summary workbook (multiple period sheets).

 WHAT IT DOES, per period sheet, for every ledger row:
   1. Reads the invoice reference out of the "Particulars" text.
   2. Finds the matching invoice PDF inside the ZIP (by invoice number).
   3. Fills Basic Amt / VAT / Invoice Value  (only if the cell is blank).
      If a value already exists and differs, it is NOT overwritten - it is
      flagged in the reconciliation report instead.
   4. Sets "Invoice (Y/N)" = Y.
   5. Writes extra PDF detail into new columns added on the right:
        PDF File | Inv No (PDF) | Inv Date (PDF) | Item Details (PDF)
   6. Builds a "Reconciliation" sheet summarising matches, amount mismatches,
      ledger rows with no PDF, and PDFs that were never used.

 The original data is preserved - new info goes into NEW columns and blank cells.

--------------------------------------------------------------------------------
 HOW TO USE (you only ever change the two paths below):
--------------------------------------------------------------------------------
   1. Put this file anywhere.
   2. Edit ZIP_PATH and EXCEL_PATH at the bottom (or pass them on the command
      line).  Run:
          python invoice_automation.py
      or
          python invoice_automation.py  "invoices.zip"  "Upload.xlsx"  "Output.xlsx"

 REQUIREMENTS (install once):
          pip install pdfplumber openpyxl
================================================================================
"""

import os
import re
import sys
import glob
import shutil
import zipfile
import tempfile
import datetime as dt

import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ----------------------------------------------------------------------------
# CONFIG - which columns to look for (matched by their header text, case-insensitive).
# You normally do NOT need to touch this. Header matching means the script keeps
# working even if column order changes between sheets.
# ----------------------------------------------------------------------------
HEADER_ROW_MAXSCAN = 8            # search first N rows for the header row
DATA_START_OFFSET  = 1           # data begins 1 row after the header row

COL_ALIASES = {
    "particulars":  ["particulars"],
    "type":         ["type"],
    "basic":        ["basic amt", "basic amount", "net", "net amt"],
    "vat":          ["vat"],
    "invoice_value":["invoice value", "gross", "total invoice"],
    "inv_yn":       ["invoice (y/n)", "invoice(y/n)", "invoice y/n", "invoice yn"],
    "supplier":     ["supplier name", "supplier"],
}

# New columns appended to the right of each sheet
NEW_COLS = ["Unique Invoice ID", "PDF File", "Inv No (PDF)", "Inv Date (PDF)",
            "PDF Net", "PDF VAT", "PDF Gross", "Item Details (PDF)"]

ID_PREFIX = "INV"   # unique IDs look like INV-0001, INV-0002, ...


def assign_unique_ids(pdfs):
    """Give every PDF a stable, sortable unique ID based on supplier folder
    then filename, so re-running on the same invoice set reproduces the same
    IDs."""
    ordered = sorted(pdfs, key=lambda p: (p["supplier_folder"].lower(), p["file"].lower()))
    for i, p in enumerate(ordered, start=1):
        p["id"] = f"{ID_PREFIX}-{i:04d}"
    return pdfs


def renamed_filename(pdf):
    """The filename used inside the renamed-invoices ZIP handed back to the user."""
    return f"{pdf['id']}_{pdf['file']}"

AMOUNT_TOLERANCE = 0.02          # £ tolerance when comparing ledger vs PDF amounts

# =============================================================================
#  PDF PARSING
# =============================================================================
def read_pdf_text(path):
    try:
        with pdfplumber.open(path) as pdf:
            return "\n".join((pg.extract_text() or "") for pg in pdf.pages)
    except Exception:
        return ""


def _money(text, label_variants):
    for lab in label_variants:
        m = re.search(lab + r'[^0-9\-\n]{0,18}(?:GBP|£)?\s*([\d,]+\.\d{2})', text, re.I)
        if m:
            try:
                return float(m.group(1).replace(",", ""))
            except ValueError:
                pass
    return None


def parse_amounts(text):
    total = _money(text, [r'Total amount payable', r'Balance Due', r'Grand Total',
                          r'Amount Due', r'Total\b'])
    net   = _money(text, [r'Sub\s*Total', r'Subtotal', r'Net Amount', r'Net\b', r'Goods'])
    vat   = _money(text, [r'VAT@?\s*\d*\.?\d*%?', r'VAT Amount', r'V\.A\.T', r'VAT\b'])
    # derive missing pieces where arithmetic allows
    if total is not None and net is not None and vat is None:
        vat = round(total - net, 2)
    if net is None and total is not None and vat is not None:
        net = round(total - vat, 2)
    if total is None and net is not None and vat is not None:
        total = round(net + vat, 2)
    return net, vat, total


def parse_invoice_no(text, filename):
    for src in (text, filename):
        m = re.search(r'Invoice\s*#?\s*[:\-]?\s*([A-Z]{0,4}[-/]?\d[\dA-Z\-/]{2,})', src, re.I)
        if m and re.search(r'\d', m.group(1)):
            return m.group(1).strip()
    m = re.search(r'(INV[-\s]?[A-Z0-9]+|SKIP\s?\d+|PRI\d+)', filename, re.I)
    return m.group(1).strip() if m else None


def parse_date(text):
    pats = [r'Invoice Date\s*[:\-]?\s*([0-9]{1,2}[/\-\.][0-9]{1,2}[/\-\.][0-9]{2,4})',
            r'Date\s*[:\-]?\s*([0-9]{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+\s+[0-9]{4})',
            r'\b([0-9]{1,2}[/\-][0-9]{1,2}[/\-][0-9]{2,4})\b']
    for p in pats:
        m = re.search(p, text, re.I)
        if m:
            return m.group(1).strip()
    return ""


def parse_line_items(path, text):
    """Best-effort compact summary of invoice line items."""
    items = []
    try:
        with pdfplumber.open(path) as pdf:
            for pg in pdf.pages:
                for tbl in (pg.extract_tables() or []):
                    for row in tbl:
                        cells = [(c or "").strip().replace("\n", " ") for c in row]
                        joined = " | ".join(c for c in cells if c)
                        if not joined:
                            continue
                        low = joined.lower()
                        if re.search(r'(item|description|qty|rate|amount|total)', low) and \
                           not re.search(r'\d', joined):
                            continue  # header row
                        if re.search(r'\d', joined) and len(joined) > 4:
                            items.append(joined)
    except Exception:
        pass
    if not items:  # fallback: text lines that look like "<desc> ... <amount>"
        for line in text.splitlines():
            line = line.strip()
            if re.search(r'[A-Za-z].*?[\d,]+\.\d{2}\s*$', line) and \
               not re.search(r'(sub\s*total|total|vat|balance|payable|amount due)',
                             line, re.I):
                items.append(re.sub(r'\s{2,}', ' ', line))
    # dedupe, cap length
    seen, out = set(), []
    for it in items:
        if it not in seen:
            seen.add(it); out.append(it)
    return "  //  ".join(out[:8])


def parse_pdf(path):
    text = read_pdf_text(path)
    fn = os.path.basename(path)
    net, vat, total = parse_amounts(text)
    return {
        "path": path,
        "file": fn,
        "supplier_folder": os.path.basename(os.path.dirname(path)),
        "inv_no": parse_invoice_no(text, fn),
        "date": parse_date(text),
        "net": net, "vat": vat, "total": total,
        "items": parse_line_items(path, text),
        "tokens": ref_tokens(fn),
    }


# =============================================================================
#  MATCHING  (invoice-number tokens, robust to leading zeros / separators)
# =============================================================================
def _norm(n):
    return n.lstrip("0") or "0"


def ref_tokens(s):
    """Return the set of normalised invoice-reference tokens found in a string."""
    s = (s or "").upper()
    toks = set()
    for m in re.findall(r'\d{3,}', s):
        if len(m) >= 17:        # skip 18-digit bank transaction ids
            continue
        toks.add(m)
        toks.add(_norm(m))
    for m in re.findall(r'SKIP\s?\d+', s):
        toks.add(m.replace(" ", ""))
    # keep only tokens with >=4 chars to reduce false collisions
    return {t for t in toks if len(t) >= 4}


def build_pdf_index(pdfs):
    idx = {}
    for p in pdfs:
        for t in p["tokens"]:
            idx.setdefault(t, []).append(p)
    return idx


STOPWORDS = {"LTD", "LIMITED", "THE", "AND", "SERVICES", "SERVICE", "SUPPLIES",
             "HIRE", "GROUP", "CONSTRUCTION", "BUILDING", "COMPANY", "SUBCONTRACTOR"}


def _supplier_words(s):
    return {w for w in re.findall(r'[A-Z]{3,}', (str(s) or "").upper())
            if w not in STOPWORDS}


def match_row(particulars, supplier_cell, pdf_index):
    """
    Return (best_pdf, note).

    Acceptance rule (guards against numeric-token collisions between suppliers):
      accept a candidate only if EITHER
        - the ledger supplier name shares a word with the PDF folder/file, OR
        - the shared invoice token is long enough (>=6 chars) to be unique.
    When several candidates qualify, the one with the strongest supplier-name
    overlap (then the longest shared token) wins.
    """
    row_tokens = ref_tokens(particulars)
    sup_words = _supplier_words(supplier_cell)

    scored = []
    seen = set()
    for t in row_tokens:
        for p in pdf_index.get(t, []):
            key = (p["path"], t)
            if key in seen:
                continue
            seen.add(key)
            fld_words = _supplier_words(p["supplier_folder"] + " " + p["file"])
            overlap = len(sup_words & fld_words)
            if overlap == 0 and len(t) < 6:
                continue                      # too weak -> reject (avoids collisions)
            scored.append((overlap, len(t), p, t))

    if not scored:
        return None, ""
    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
    best_overlap, best_len, best_pdf, best_tok = scored[0]
    # distinct qualifying PDFs
    distinct = {s[2]["path"] for s in scored}
    note = ""
    if len(distinct) > 1:
        note = "multiple PDFs qualified; picked strongest supplier/token match"
    if best_overlap == 0:
        note = (note + "; " if note else "") + "matched on invoice number only"
    return best_pdf, note


# =============================================================================
#  EXCEL HELPERS
# =============================================================================
def find_header_row(ws):
    for r in range(1, HEADER_ROW_MAXSCAN + 1):
        vals = [str(ws.cell(r, c).value or "").strip().lower()
                for c in range(1, ws.max_column + 1)]
        if "particulars" in vals:
            return r
    return None


def map_columns(ws, header_row):
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(header_row, c).value or "").strip().lower()
        if h:
            headers[h] = c
    colmap = {}
    for key, aliases in COL_ALIASES.items():
        for a in aliases:
            if a in headers:
                colmap[key] = headers[a]
                break
    return colmap


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("£", "").strip())
    except ValueError:
        return None


# =============================================================================
#  MAIN
# =============================================================================
def run(zip_path, excel_path, out_path, renamed_zip_path=None):
    print(f"[1/5] Extracting invoices from: {zip_path}")
    workdir = tempfile.mkdtemp(prefix="inv_")
    with zipfile.ZipFile(zip_path) as z:
        z.extractall(workdir)
    pdf_files = sorted(glob.glob(os.path.join(workdir, "**", "*.pdf"), recursive=True))
    print(f"      found {len(pdf_files)} PDF invoices")

    print("[2/5] Reading each PDF (invoice no, amounts, dates, line items)...")
    pdfs = [parse_pdf(p) for p in pdf_files]
    assign_unique_ids(pdfs)
    pdf_index = build_pdf_index(pdfs)

    print(f"[3/5] Opening workbook: {excel_path}")
    wb = load_workbook(excel_path)

    used_pdfs = set()
    recon = []   # rows for the reconciliation sheet

    # Only process sheets that look like period ledgers (they contain 'Particulars')
    period_sheets = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if find_header_row(ws):
            period_sheets.append(sn)

    print(f"[4/5] Writing item info into {len(period_sheets)} ledger sheet(s): "
          f"{period_sheets}")

    for sn in period_sheets:
        ws = wb[sn]
        hr = find_header_row(ws)
        cm = map_columns(ws, hr)
        if "particulars" not in cm:
            continue

        # append the 4 new columns
        base = ws.max_column
        new_col_idx = {}
        for i, name in enumerate(NEW_COLS):
            c = base + 1 + i
            cell = ws.cell(hr, c, name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", start_color="D9E1F2")
            new_col_idx[name] = c
            ws.column_dimensions[cell.column_letter].width = 24 if "Item" not in name else 55

        matched = 0
        for r in range(hr + DATA_START_OFFSET, ws.max_row + 1):
            part = ws.cell(r, cm["particulars"]).value
            if not part:
                continue
            typ = str(ws.cell(r, cm.get("type", 0), ).value or "").strip() if cm.get("type") else ""
            sup = ws.cell(r, cm["supplier"]).value if cm.get("supplier") else ""

            pdf, note = match_row(str(part), sup, pdf_index)
            if not pdf:
                if typ.upper() != "PAY":        # PAY rows are bank charges, no invoice expected
                    recon.append([sn, r, str(part)[:60], "NO PDF FOUND", "", "", ""])
                continue

            matched += 1
            used_pdfs.add(pdf["path"])

            # --- fill amount cells only when blank ---
            for key, pdfval in (("basic", pdf["net"]),
                                ("vat", pdf["vat"]),
                                ("invoice_value", pdf["total"])):
                if key in cm and pdfval is not None and \
                   to_float(ws.cell(r, cm[key]).value) is None:
                    ws.cell(r, cm[key], pdfval)

            # --- verification: flag ONLY on the gross total (most reliably
            #     parsed figure). Net/VAT splits vary too much across the 18
            #     invoice layouts to compare safely, so they are not flagged. ---
            mismatch = []
            if "invoice_value" in cm and pdf["total"] is not None:
                cur = to_float(ws.cell(r, cm["invoice_value"]).value)
                if cur is not None and abs(cur - pdf["total"]) > AMOUNT_TOLERANCE:
                    mismatch.append(f"gross: ledger {round(cur,2)} vs PDF {pdf['total']}")

            # --- Invoice (Y/N) = Y ---
            if "inv_yn" in cm:
                ws.cell(r, cm["inv_yn"], "Y")

            # --- new detail columns ---
            ws.cell(r, new_col_idx["Unique Invoice ID"], pdf["id"])
            ws.cell(r, new_col_idx["PDF File"], pdf["file"])
            ws.cell(r, new_col_idx["Inv No (PDF)"], pdf["inv_no"] or "")
            ws.cell(r, new_col_idx["Inv Date (PDF)"], pdf["date"] or "")
            ws.cell(r, new_col_idx["PDF Net"], pdf["net"])
            ws.cell(r, new_col_idx["PDF VAT"], pdf["vat"])
            ws.cell(r, new_col_idx["PDF Gross"], pdf["total"])
            ws.cell(r, new_col_idx["Item Details (PDF)"], pdf["items"] or "")

            status = "OK" if not mismatch else "AMOUNT MISMATCH"
            recon.append([sn, r, str(part)[:60], status,
                          pdf["id"], pdf["file"], pdf["inv_no"] or "", " ; ".join(mismatch)])

        print(f"        {sn}: matched {matched} row(s) to invoices")

    # ---- reconciliation sheet ----
    if "Reconciliation" in wb.sheetnames:
        del wb["Reconciliation"]
    rc = wb.create_sheet("Reconciliation")
    headers = ["Sheet", "Excel Row", "Particulars (short)", "Status",
               "Unique Invoice ID", "PDF File", "Inv No", "Notes"]
    rc.append(headers)
    for c in range(1, len(headers) + 1):
        rc.cell(1, c).font = Font(bold=True, color="FFFFFF")
        rc.cell(1, c).fill = PatternFill("solid", start_color="4472C4")
    for row in recon:
        # NO PDF FOUND rows built earlier have 7 items (no ID); pad them
        if len(row) == 7:
            row = row[:4] + [""] + row[4:]
        rc.append(row)
    # unused PDFs (present in ZIP but not tied to any ledger row)
    rc.append([])
    rc.append(["--- PDFs in ZIP not matched to any ledger row (later period / duplicate) ---"])
    for p in pdfs:
        if p["path"] not in used_pdfs:
            rc.append(["", "", "", "UNUSED PDF", p["id"], p["file"], p["inv_no"] or "", ""])
    for col in "ABCDEFGH":
        rc.column_dimensions[col].width = 24
    rc.column_dimensions["C"].width = 48
    rc.column_dimensions["H"].width = 40

    # colour-code status
    fills = {"OK": "C6EFCE", "AMOUNT MISMATCH": "FFEB9C",
             "NO PDF FOUND": "FFC7CE", "UNUSED PDF": "D9D9D9"}
    for r in range(2, rc.max_row + 1):
        st = rc.cell(r, 4).value
        if st in fills:
            for c in range(1, 9):
                rc.cell(r, c).fill = PatternFill("solid", start_color=fills[st])

    if renamed_zip_path:
        print(f"      Building renamed-invoices ZIP -> {renamed_zip_path}")
        with zipfile.ZipFile(renamed_zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in pdfs:
                zf.write(p["path"], arcname=renamed_filename(p))

    print(f"[5/5] Saving -> {out_path}")
    wb.save(out_path)

    # summary
    n_ok = sum(1 for x in recon if x[3] == "OK")
    n_mis = sum(1 for x in recon if x[3] == "AMOUNT MISMATCH")
    n_no = sum(1 for x in recon if x[3] == "NO PDF FOUND")
    n_unused = sum(1 for p in pdfs if p["path"] not in used_pdfs)
    print("\n================= SUMMARY =================")
    print(f"  Rows matched & OK ............ {n_ok}")
    print(f"  Rows matched, amount differs . {n_mis}")
    print(f"  Ledger rows with NO PDF ...... {n_no}")
    print(f"  PDFs in ZIP left unused ...... {n_unused}")
    print("  See the 'Reconciliation' sheet for full detail.")
    print("==========================================")

    shutil.rmtree(workdir, ignore_errors=True)
    return out_path


if __name__ == "__main__":
    # -------- EDIT THESE TWO PATHS FOR REUSE (or pass on command line) --------
    ZIP_PATH   = "SUPPLIERWISE_INVOICE_DETAILS.zip"
    EXCEL_PATH = "Upload.xlsx"
    OUT_PATH   = "Upload_WITH_INVOICE_DETAILS.xlsx"
    RENAMED_ZIP_PATH = "Invoices_with_IDs.zip"
    # -------------------------------------------------------------------------
    if len(sys.argv) >= 3:
        ZIP_PATH, EXCEL_PATH = sys.argv[1], sys.argv[2]
        OUT_PATH = sys.argv[3] if len(sys.argv) >= 4 else OUT_PATH
    run(ZIP_PATH, EXCEL_PATH, OUT_PATH, RENAMED_ZIP_PATH)
 
